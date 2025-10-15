import pathlib
from class_StaV import Sta_V
from StaticVar import StaticVar as persistent

import configparser
import os
import operator
from distance import distance
from importdata import importdata
import matplotlib
import matplotlib.pyplot as plt
import openpyxl
from EEW_Single import EEW_Single
from function1 import *
import time as Time

startT2 = 0

matplotlib.use('TkAgg')  # 或者 'Qt5Agg', 'PyQt5', 'tk' 等
from openpyxl import load_workbook


def EEW_Test_Massvie(datadirRRR, excelName, Num):
    min_time = startT2
    time_start = Time.strftime("%Y-%m-%d %H:%M:%S", Time.localtime())
    # if persistent.StartT1 < 0:
    #     persistent.StartT1 = min_time
    # if persistent.position < 0:
    #     persistent.position = 0
    # cnt=counter()
    EEW_Params = configparser.ConfigParser()
    EEW_Params.read(os.path.join(os.getcwd(), 'EEW_Params.ini'), encoding='UTF-8')
    Debug = EEW_Params['EEW_Single']['Debug']  # 字符非int
    Debug = int(Debug)
    # Debug =1
    deltSecond = 0.1
    directory = os.listdir(datadirRRR)
    filename1 = directory[0]
    Nfile = len(directory)
    # print(filename1[-1])
    if (operator.eq(filename1[-1], '1')) == 1:
        flag1 = 1
    else:
        flag1 = 0

    fileToread = []
    info_all = []  #
    for k in range(0, Nfile, 3 * (flag1 + 1)):
        E = -1
        N = -1
        Z = -1
        E1 = -1
        N1 = -1
        Z1 = -1
        E2 = -1
        N2 = -1
        Z2 = -1
        E11 = -1
        E22 = -1
        N11 = -1
        N22 = -1
        Z11 = -1
        Z22 = -1
        if flag1 == 0:
            filename1 = directory[k]
            filename2 = directory[k + 1]
            filename3 = directory[k + 2]
            if str.find(filename1, 'EW') < 0 and str.find(filename1, 'ew') < 0:
                print('不是你以为的EW')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename1)
            if str.find(filename2, 'NS') < 0 and str.find(filename2, 'ns') < 0:
                print('不是你以为的NS')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename2)
            if str.find(filename3, 'UD') < 0 and str.find(filename3, 'ud') < 0:
                print('不是你以为的UD', 'UD')
            fileToread.append(datadirRRR + '\\' + filename3)
        elif flag1 == 1:

            filename1 = directory[k]
            filename2 = directory[k + 1]
            filename3 = directory[k + 2]
            filename4 = directory[k + 3]
            filename5 = directory[k + 4]
            filename6 = directory[k + 5]

            if str.find(filename1, 'EW1') < 0 and str.find(filename1, 'ew1') < 0:
                print('不是你以为的EW1')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename1)
            if str.find(filename2, 'EW2') < 0 and str.find(filename2, 'ew2') < 0:
                print('不是你以为的EW2')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename2)
            if str.find(filename3, 'NS1') < 0 and str.find(filename3, 'ns1') < 0:
                print('不是你以为的NS1')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename3)
            if str.find(filename4, 'NS2') < 0 and str.find(filename4, 'ns2') < 0:
                print('不是你以为的NS2')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename4)
            if str.find(filename5, 'UD1') < 0 and str.find(filename5, 'ud1') < 0:
                print('不是你以为的UD1')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename5)
            if str.find(filename6, 'UD2') < 0 and str.find(filename6, 'ud2') < 0:
                print('不是你以为的UD2')
                os.system('pause')
            fileToread.append(datadirRRR + '\\' + filename6)

        # 文件读取
        file_obj = open(fileToread[k])
        txt = file_obj.readlines()
        cont1 = txt[13].find("(")
        cont2 = txt[13].rfind("/")
        p1 = float(txt[13][18:cont1])
        p2 = float(txt[13][cont2 + 1:])
        Factor = p1 / p2  # 因子
        ind = 17
        E1 = importdata(txt, ind)
        E1 = np.array(E1, dtype=float) * Factor  # np.array(E1) * Factor
        # E1 = E1
        file_obj.close()
        filecount = 3 * (flag1 + 1)

        if filecount == 3:  # 三通道传感器
            file_obj = open(fileToread[k + 1])
            txt = file_obj.readlines()
            N1 = importdata(txt, ind)
            N1 = np.array(N1) * Factor
            N1 = N1
            file_obj.close()
            file_obj = open(fileToread[k + 2])
            txt = file_obj.readlines()
            Z1 = importdata(txt, ind)
            Z1 = np.array(Z1) * Factor
            Z1 = Z1
            file_obj.close()
            E2 = E1
            N2 = N1
            Z2 = Z1
        else:  # 6通道传感器
            file_obj = open(fileToread[k + 1])
            txt = file_obj.readlines()
            E2 = importdata(txt, ind)
            E2 = np.array(E2) * Factor
            E2 = E2
            file_obj.close()
            file_obj = open(fileToread[k + 2])
            txt = file_obj.readlines()
            N1 = importdata(txt, ind)
            N1 = np.array(N1) * Factor
            N1 = N1
            file_obj.close()
            file_obj = open(fileToread[k + 3])
            txt = file_obj.readlines()
            N2 = importdata(txt, ind)
            N2 = np.array(N2) * Factor
            N2 = np.array(N2)
            file_obj.close()
            file_obj = open(fileToread[k + 4])
            txt = file_obj.readlines()
            Z1 = importdata(txt, ind)
            Z1 = np.array(Z1) * Factor
            Z1 = Z1
            file_obj.close()
            file_obj = open(fileToread[k + 5])
            txt = file_obj.readlines()
            Z2 = importdata(txt, ind)
            Z2 = np.array(Z2) * Factor
            Z2 = Z2
            file_obj.close()
        sprate = int(txt[10][18:21])
        StCode = txt[5][18:-1]
        epiLat = float(txt[1][18:-1])
        epiLon = float(txt[2][18:-1])
        Lat = float(txt[6][18:-1])
        Lon = float(txt[7][18:-1])
        magnitude = float(txt[4][18:-1])
        Azimuth_real = getDegree(Lat, Lon, epiLat, epiLon)
        epi_dist = distance(epiLat, epiLon, Lat, Lon)
        EQ_Info = filename1[0:-3]
        enz1 = E1 * E1 + N1 * N1 + Z1 * Z1
        squares1 = [math.sqrt(num) for num in enz1]
        PGA_Real1 = max(squares1)
        enz2 = E2 * E2 + N2 * N2 + Z2 * Z2
        squares2 = [math.sqrt(num) for num in enz2]
        PGA_Real2 = max(squares2)
        trace1 = [E1, N1, Z1, E2, N2, Z2]
        tempTime0 = txt[9][18:-1].split("\t")
        tempTime = tempTime0[0].split(" ")
        tempTime1 = tempTime[0]
        tempTime2 = tempTime[1]
        tempTime3 = (tempTime1 + ' ' + tempTime2)

        if len(tempTime[0]) > 8:
            tempTime2 = tempTime[1]
            time_parts = tempTime2.split(':')
            if len(time_parts) >= 2:
                minutes = int(time_parts[1])
                if minutes > 59:  # 如果分钟数无效，修正为 59
                    time_parts[1] = '59'
                    tempTime2 = ':'.join(time_parts)
            tempTime3 = (tempTime1 + ' ' + tempTime2)
        try:
            ans = Time.strptime(tempTime3, '%Y/%m/%d %H:%M:%S.%f')  # time objecct
        except:
            ans = Time.strptime(tempTime3, '%Y/%m/%d %H:%M:%S')  # time objecct
        else:
            time_parts = tempTime2.split(':')
            tempTime2 = tempTime[1]
            if len(time_parts) >= 2:
                minutes = int(time_parts[1])
                if minutes > 59:  # 如果分钟数无效，修正为 59
                    time_parts[1] = '59'
                    tempTime2 = ':'.join(time_parts)
            tempTime1 =   tempTime[0]    # '20' + tempTime[0]
            # tempTime3 = (tempTime1 + ' ' + tempTime2)
            try:
                ans = Time.strptime(tempTime3, '%Y/%m/%d %H:%M:%S.%f')
            except:
                ans = Time.strptime(tempTime3, '%Y/%m/%d %H:%M:%S')

        if ans is None:
            ans = 0
        record_time = Time.mktime(ans)  # timestamp format：unix time

        epi_time = record_time + 10  # 暂时加上
        ## 上采样、降采样模块
        ##
        tadd = 30  # 秒
        temp_traceE1 = np.tile(trace1[0][0:200], tadd)
        temp_traceN1 = np.tile(trace1[1][0:200], tadd)
        temp_traceZ1 = np.tile(trace1[2][0:200], tadd)
        temp_traceE2 = np.tile(trace1[3][0:200], tadd)
        temp_traceN2 = np.tile(trace1[4][0:200], tadd)
        temp_traceZ2 = np.tile(trace1[5][0:200], tadd)
        E11 = np.append(temp_traceE1, E1)
        N11 = np.append(temp_traceN1, N1)
        Z11 = np.append(temp_traceZ1, Z1)
        E22 = np.append(temp_traceE2, E2)
        N22 = np.append(temp_traceN2, N2)
        Z22 = np.append(temp_traceZ2, Z2)
        trace2 = np.array([E11, N11, Z11, E22, N22, Z22])
        trace2 = np.transpose(trace2)

        Sta_vars1 = Sta_V()  # 对象初始化
        Sta_vars2 = Sta_V()  # 对象初始化
        Sta_vars1.STA_Name = StCode
        Sta_vars1.STA_Long = Lon
        Sta_vars1.STA_Lat = Lat
        Sta_vars1.sprate = sprate
        Sta_vars1.PGA_Real = PGA_Real1
        Sta_vars1.Distance_real = epi_dist
        Sta_vars1.Magnitude_real = magnitude
        Sta_vars1.Epi_time_real = epi_time
        Sta_vars1.Aziumth_real = Azimuth_real
        Sta_vars1.epiLat_real = epiLat
        Sta_vars1.epiLon_real = epiLon

        Sta_vars2.STA_Name = StCode
        Sta_vars2.STA_Long = Lon
        Sta_vars2.STA_Lat = Lat
        Sta_vars2.sprate = sprate
        Sta_vars2.PGA_Real = PGA_Real1
        Sta_vars2.Distance_real = epi_dist
        Sta_vars2.Magnitude_real = magnitude
        Sta_vars2.Epi_time_real = epi_time
        Sta_vars2.Aziumth_real = Azimuth_real
        Sta_vars2.epiLat_real = epiLat
        Sta_vars2.epiLon_real = epiLon
        delt = int(float(deltSecond) * sprate)
        L = int(np.fix(len(E11) / delt))
        # try:
        #     persistent.StarT
        # except:
        #     os.system("pause")
        if Num == 0 or persistent.StartT < 0:
            persistent.StartT = 0.005

            everyStartT = persistent.StartT

        info_real = [time_start, Num]
        plt.ion()
        ax = []
        ay = []
        bx = []
        by = []
        sst = []
        st0 = startT2
        print(st0)
        for ttt in range(L - 1):
            # if math.ceil(persistent.StartT-everyStartT)*sprate+delt>len(trace2):
            #     break
            # datak=trace2
            #
            # if persistent.StartT==0.005:
            #     Data_now=datak[0:60*sprate]
            #     StartT=persistent.StartT
            #     [Sta_vars11,Sta_vars22]=EEW_Single(Data_now, StartT,Lon,Lat,Sta_vars1,Sta_vars2)
            #     persistent.StartT=persistent.StartT+60
            # else:
            #    d1=math.ceil(persistent.StartT-everyStartT)*sprate
            #    d2=math.ceil(persistent.StartT-everyStartT)*sprate+delt
            #     # print(Data_now)
            #    Data_now[d1:d2,:]
            datak = trace2
            num = (ttt) * delt
            Data_now = datak[num:delt * (ttt + 1)]
            StartT = np.round(min_time + (ttt) * delt / sprate, 3)
            qq = np.round(delt / sprate, 3)
            increment_T(qq)
            t = (ttt) * delt / sprate  # 画图专用time
            iskkk = -1
            if len(Data_now) == 0 or StartT == None:
                print("EEEW_Single输入为空，请检查！！！")
            # [NewInfo, StationInfosStruct,Sta_vars_ret,StartT1,Sta_vars_ret2,AlarmS]=EEW_Single(Data_now, StartT,iskkk,Lon,Lat,Sta_vars1,Sta_vars2)‘
            start_time = Time.time()
            # [NewInfo, StationInfo, Sta_vars_ret1, Sta_vars_ret2, StartT, AlarmS] = EEW_Single(Data_now, StartT, Lon, Lat, Sta_vars1, Sta_vars2)
            # [NewInfo, StationInfo, Sta_vars_ret1, Sta_vars_ret2, StartT, AlarmS] = EEW_Single(Data_now,startT2,magnitude,Num)

            # try:
            [NewInfo, StationInfo, Sta_vars_ret1, Sta_vars_ret2, StartT, AlarmS] = EEW_Single(Data_now, startT2)
            sst.append(startT2)
            # except Exception as e:
            #     print(e)
            #     # breakpoint()
            #     os.system("pasue")

            # plot debuge=2时才启动
            if Debug == 2:
                # if t>=68.5:
                #     print('暂停观察')
                Data_now1 = Data_now[:, 2]
                subplot22(delt, Sta_vars_ret1, Sta_vars_ret2, sprate, len(Data_now1), min_time)

            # end_time = Time.time()
            # elapsed_time = end_time - start_time
            # # print(elapsed_time)
            try:
                NewInfo
            except Exception as e1:
                print(e1)
            if NewInfo == 4 or NewInfo == 100:
                print(AlarmS)
            if NewInfo > 0:  # NewInfo == 1 or NewInfo == 3 or NewInfo == 2:
                # EpLon_pre = StationInfo.Epi_Long
                # EpLat_pre = StationInfo.Epi_Lat
                EpLon_pre, EpLat_pre = LonLat(Lon, Lat, StationInfo.Azimuth, StationInfo.Distance)
                StationInfo.Epi_Long = EpLon_pre
                StationInfo.Epi_Lat = EpLat_pre
                pi = math.pi
                delt_Epi = 6371.004 * math.acos((math.sin(EpLat_pre / 180 * pi) * math.sin(
                    epiLat / 180 * pi) + math.cos(EpLat_pre / 180 * pi) * math.cos(epiLat / 180 * pi) * math.cos(
                    (epiLon - EpLon_pre) / 180 * pi)))
                Pose = (StationInfo.P_time - tadd - min_time) * sprate
                info_one = (
                time_start, Num, EQ_Info, StCode, epi_time, epiLon, epiLat, magnitude, epi_dist, sprate, Lon, Lat,
                PGA_Real1, Azimuth_real, StartT,
                StationInfo.P_time, StationInfo.S_time, StationInfo.Magnitude, StationInfo.Azimuth,
                StationInfo.Epi_time, StationInfo.Distance,
                StationInfo.Epi_Long, StationInfo.Epi_Lat, StationInfo.PGA_Pred, StationInfo.S_time_cal,
                StationInfo.PGA_Curr, StationInfo.PGV_Curr, StationInfo.PGD_Curr
                , StationInfo.DurCurr, StationInfo.S_time, Sta_vars2.AlarmLevel, delt_Epi, Pose,
                NewInfo, StationInfo.PGA_Pred)
                info_all.append(info_one)
        plt.ioff()
        # 无报警存一行基础信息
        if len(info_all) == 0 or StCode != info_all[-1][3]:
            info_one = (
                time_start, Num, EQ_Info, StCode, epi_time, epiLon, epiLat, magnitude, epi_dist, sprate, Lon, Lat,
                PGA_Real1, Azimuth_real, 0, 0, 0, 0, 0, 0, 0,
                0, 0, 0, 0, 0, 0, 0 , 0, 0, 0, 0, 0, 0, 0)
            info_all.append(info_one)
        # # save figure
        psall = np.empty((len(info_all), 2))
        for ss in range(len(info_all)):
            psall[ss, 0] = info_all[ss][15]
            psall[ss, 1] = info_all[ss][16]
        savefigure(trace2.T, sprate, psall, len(Data_now[:, 2]), min_time, st0, EQ_Info)


    # excel 操作,最后一次性写
    ws_name = 'Test_result'
    # info_all1=list(info_all)

    if not pathlib.Path(excelName).exists():
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(excelName)

    # ws=ws[ws_name]
    heads = ['time_start', 'epi_No.', 'EQ_Info', 'STA_Name', 'Epi_time', 'epiLon_real', 'epiLat_real', 'Magnitude_real',
             'Distance_Real', 'sprate',
             'STA_Long', 'STA_Lat', 'PGA_Real', 'Azimuth_real', 'StartT', 'P_time', 'S_time', 'Magnitude', 'Azimuth',
             'Epi_time',
             'Distance', 'Epi_Long', 'Epi_Lat', 'PGA_Pred', 'S_time_cal', 'PGA_Curr', 'PGV_Curr', 'PGD_Curr', 'DurCurr',
             'S_time2', 'AlarmLevel', 'delt_Epi', 'POS', 'Newinfo', 'PGAPre']
    ws = wb.active
    Title_len = len(heads)
    max_len = ws.max_row
    title_col = 1
    if max_len == 1:
        write_row = 1
    else:
        write_row = max_len + 1

    for i in range(Title_len):
        ws.cell(row=write_row, column=title_col, value=heads[i])
        title_col += 1
    mrow = 0
    if info_all:
        L = len(info_all)
        LL = np.shape(info_all)[1]
    else:
        L == 0
        LL = 0
    Sta_vars1

    countrows = 0
    max_len = ws.max_row
    # deltline=
    for rows in range(max_len + 1, max_len + L + 1):

        for line in range(1, LL + 1):
            ws.cell(row=rows, column=line, value=info_all[countrows][line - 1])
        countrows = countrows + 1

        mrow = mrow + 1
        mline = 0

        # datak=trace2
        # StartT=[Num*10+(ttt-1)*delt/sprate+1
        # [Sta_vars11,Sta_vars22]=EEW_Single(Data_now, StartT,Lon,Lat,Sta_vars1,Sta_vars2)

        # Sta_vars1=Sta_vars11
        # Sta_vars2=Sta_vars22

    wb.save(excelName)
    # os.system("pause")

    # sst
    with open('tttt.txt', 'a+') as file:
        file.write(str(sst))


# os.system("pause")
def getDegree(latA, lonA, latB, lonB):
    """
    Args:
      point p1(latA, lonA)
      point p2(latB, lonB)
    Returns:
      bearing between the two GPS points,
      default: the basis of heading direction is north
    """
    radLatA = math.radians(latA)
    radLonA = math.radians(lonA)
    radLatB = math.radians(latB)
    radLonB = math.radians(lonB)
    dLon = radLonB - radLonA
    y = math.sin(dLon) * math.cos(radLatB)
    x = math.cos(radLatA) * math.sin(radLatB) - math.sin(radLatA) * math.cos(radLatB) * math.cos(dLon)
    brng = math.degrees(math.atan2(y, x))
    brng = (brng + 360) % 360
    return brng


def subplot22(delt, Sta_vars_ret1, Sta_vars_ret2, Sprate, packageLen, min_time):
    # delt:发包的长度
    # ax:list,子图1的横坐标
    # ay:list,组图1的纵坐标
    # bx:list,子图2的横坐标
    # by:list,组图2的纵坐标
    # Sta_vars_ret1:提取所需信息
    # Sprate：采样率
    # packageLen:最新包长
    # min_time:时间归一起点
    sensordata1 = Sta_vars_ret1.Buffer  # Buffer
    sensordata2 = Sta_vars_ret2.Buffer  # Buffer
    n1 = len(sensordata1[2])
    st0 = Sta_vars_ret1.StartT + (packageLen - n1) / Sprate

    tout = cmp_t(n1, Sprate, st0 - min_time)
    ax = tout.tolist()
    bx = tout.tolist()
    ay = sensordata1[2]
    by = sensordata2[2]
    ptime = Sta_vars_ret1.P_time
    TrigPN = int((Sta_vars_ret1.P_time - st0) * Sprate)  # POINTS
    stime = Sta_vars_ret1.S_time
    TrigSN = int((Sta_vars_ret1.S_time - st0) * Sprate)  # POINTS
    endtime = Sta_vars_ret1.End_time
    TrigEN = int((Sta_vars_ret1.End_time - st0) * Sprate)  # POINTS
    Is_EQK = Sta_vars_ret1.Is_EQK

    # plt.ion()    # 开启一个画图的窗口进入交互模式，用于实时更新数据
    # plt.rcParams['savefig.dpi'] = 200 #图片像素
    # plt.rcParams['figure.dpi'] = 200 #分辨率
    plt.rcParams['figure.figsize'] = (5, 3.5)  # 图像显示大小
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 防止中文标签乱码，还有通过导入字体文件的方法
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['lines.linewidth'] = 0.5  # 设置曲线线条宽度
    plt.clf()  # 清除刷新前的图表，防止数据量过大消耗内存
    plt.suptitle("实时波形")  # , fontsize=5, 添加总标题，并设置文字大小

    # 图表1
    agraphic = plt.subplot(2, 1, 1)
    agraphic.set_title('传感器1_未去基线')  # 添加子标题
    agraphic.set_xlabel('x轴')  # , fontsize=5, 添加轴标签
    agraphic.set_ylabel('y轴')  # , fontsize=5
    agraphic.plot(ax, ay, 'b-')  # 等于agraghic.plot(ax,ay,'g-')
    if TrigPN > 0 and TrigPN < n1:
        agraphic.plot(ax[TrigPN], ay[TrigPN], 'r.', linewidth=2)  # 触发位置加入图示中
    if TrigSN > 0 and TrigSN < n1:
        agraphic.plot(ax[TrigSN], ay[TrigSN], 'g.', linewidth=2)  # S触发位置加入图示中
    if TrigEN > 0 and TrigEN < n1:
        agraphic.plot(ax[TrigEN], ay[TrigEN], 'k.', linewidth=2)  # end位置加入图示中

    # 图表2
    bgraghic = plt.subplot(2, 1, 2)
    bgraghic.plot(bx, by, 'b-')
    bgraghic.set_xlabel('x轴')  # , fontsize=5, 添加轴标签
    bgraghic.set_ylabel('y轴')  # , fontsize=5
    if TrigPN > 0 and TrigPN < n1:
        bgraghic.plot(ax[TrigPN], ay[TrigPN], 'r.', linewidth=2)  # 触发位置加入图示中
    if TrigSN > 0 and TrigSN < n1:
        bgraghic.plot(ax[TrigSN], ay[TrigSN], 'g.', linewidth=2)  # S触发位置加入图示中
    if TrigEN > 0 and TrigEN < n1:
        bgraghic.plot(ax[TrigEN], ay[TrigEN], 'k.', linewidth=4)  # end位置加入图示中
    bgraghic.set_title('传感器2_未去基线')
    plt.pause(0.001)  # 设置暂停时间，太快图表无法正常显示


def increment_T(x):
    global startT2  # 声明全局变量
    startT2 = np.round(startT2 + x, 3)  # 自加操作x


def savefigure(trace2, Sprate, P_S, packageLen, min_time, st0, EQ_Info):
    # delt:发包的长度
    # ax:list,子图1的横坐标
    # ay:list,组图1的纵坐标
    # bx:list,子图2的横坐标
    # by:list,组图2的纵坐标
    # trace1:原始数据，行
    # trace2:数据，add head,行
    # P_S：ptime\Stime,
    # packageLen:最新包长
    # min_time:时间归一起点
    n1 = len(trace2[0])
    # st0 = 0  # Sta_vars_ret1.StartT + (packageLen - n1) / Sprate
    # 传感器1
    tout = cmp_t(n1, Sprate, 0)
    ax = tout.tolist()
    bx = tout.tolist()
    cx = tout.tolist()
    ay = trace2[2]
    by = trace2[0]
    cy = trace2[1]

    pall = P_S[:, 0]
    # pall=pall(pall>0)
    pall = [element for element in pall if element > 0]
    unique_values = np.unique(pall)
    ppos = [np.where(pall == val)[0] for val in unique_values]  # ppos[i][0]
    sall = P_S[:, 1]
    # sall = sall[sall > 0]
    sall = [element for element in sall if element > 0]
    unique_values1 = np.unique(sall)
    spos = [np.where(sall == val)[0] for val in unique_values1]

    TrigPN = np.empty((len(ppos), 1))
    if len(ppos) > 0:
        for i in range(len(ppos)):
            ptime = pall[ppos[i][0]]
            TrigPN[i, :] = int((ptime - st0) * Sprate)  # POINTS
    TrigPN = [element for element in TrigPN if element > 0]
    TrigPN = np.array(TrigPN)
    TrigPN = TrigPN.flatten().astype(int)
    TrigSN = np.empty((len(spos), 1))
    if len(spos) > 0:
        for i in range(len(spos)):
            stime = sall[spos[i][0]]
            TrigSN[i, :] = int((stime - st0) * Sprate)  # POINTS
    TrigSN = [element for element in TrigSN if element > 0]
    TrigSN = np.array(TrigSN)
    TrigSN = TrigSN.flatten().astype(int)

    # Sta_vars_ret1中没有结束时间和Is_EQK，暂无法标记
    # endtime = Sta_vars_ret1.End_time
    # TrigEN = int((Sta_vars_ret1.End_time - st0) * Sprate)  # POINTS
    Is_EQK = 1  # Sta_vars_ret1.Is_EQK

    # plt.ion()    # 开启一个画图的窗口进入交互模式，用于实时更新数据
    # plt.rcParams['savefig.dpi'] = 200 #图片像素
    # plt.rcParams['figure.dpi'] = 200 #分辨率
    plt.rcParams['figure.figsize'] = (10, 7)  # 图像显示大小
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 防止中文标签乱码，还有通过导入字体文件的方法
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['lines.linewidth'] = 0.5  # 设置曲线线条宽度
    plt.clf()  # 清除刷新前的图表，防止数据量过大消耗内存
    plt.suptitle(EQ_Info)  # , fontsize=5, 添加总标题，并设置文字大小

    # 图表1
    agraphic = plt.subplot(3, 1, 1)
    # agraphic.set_title('传感器1_Z')  # 添加子标题
    agraphic.set_xlabel('time(s)')  # , fontsize=5, 添加轴标签
    agraphic.set_ylabel('Z')  # , fontsize=5
    agraphic.plot(ax, ay, 'b-')  # 等于agraghic.plot(ax,ay,'g-')
    if len(TrigPN) > 0 and TrigPN[-1] < n1:
        # agraphic.plot(ax[TrigPN], ay[TrigPN], 'r.', linewidth=2)  # 触发位置加入图示中
        for i in range(len(TrigPN)):
            plt.axvline(ax[TrigPN[i]], color='r', linestyle='--')
    if len(TrigSN) > 0 and TrigSN[-1] < n1:
        # agraphic.plot(ax[TrigSN], ay[TrigSN], 'g.', linewidth=2)  # S触发位置加入图示中
        for i in range(len(TrigSN)):
            plt.axvline(ax[TrigSN[i]], color='g', linestyle='-')
        # if TrigEN > 0 and TrigEN < n1:
    #     agraphic.plot(ax[TrigEN], ay[TrigEN], 'k.', linewidth=2)  # end位置加入图示中

    # 图表2
    bgraghic = plt.subplot(3, 1, 2)
    bgraghic.plot(bx, by, 'b-')
    bgraghic.set_xlabel('time(s)')  # , fontsize=5, 添加轴标签
    bgraghic.set_ylabel('E')  # , fontsize=5
    if len(TrigPN) > 0 and TrigPN[-1] < n1:
        for i in range(len(TrigPN)):
            # bgraghic.plot(ax[TrigPN], ay[TrigPN], 'r.', linewidth=2)  # 触发位置加入图示中
            plt.axvline(ax[TrigPN[i]], color='r', linestyle='--')
    if len(TrigSN) > 0 and TrigSN[-1] < n1:
        for i in range(len(TrigSN)):
            # bgraghic.plot(ax[TrigSN], ay[TrigSN], 'g.', linewidth=2)  # S触发位置加入图示中
            plt.axvline(ax[TrigSN[i]], color='g', linestyle='-')
    # if TrigEN > 0 and TrigEN < n1:
    #     bgraghic.plot(ax[TrigEN], ay[TrigEN], 'k.', linewidth=4)  # end位置加入图示中
    # bgraghic.set_title('传感器2_未去基线')

    # 图表3
    bgraghic = plt.subplot(3, 1, 3)
    bgraghic.plot(cx, cy, 'b-')
    bgraghic.set_xlabel('time(s)')  # , fontsize=5, 添加轴标签
    bgraghic.set_ylabel('N')  # , fontsize=5
    if len(TrigPN) > 0 and TrigPN[-1] < n1:
        for i in range(len(TrigPN)):
            # bgraghic.plot(ax[TrigPN], ay[TrigPN], 'r.', linewidth=2)  # 触发位置加入图示中
            plt.axvline(ax[TrigPN[i]], color='r', linestyle='--')
    if len(TrigSN) > 0 and TrigSN[-1] < n1:
        for i in range(len(TrigSN)):
            # bgraghic.plot(ax[TrigSN], ay[TrigSN], 'g.', linewidth=2)  # S触发位置加入图示中
            plt.axvline(ax[TrigSN[i]], color='g', linestyle='-')
    # if TrigEN > 0 and TrigEN < n1:
    #     bgraghic.plot(ax[TrigEN], ay[TrigEN], 'k.', linewidth=4)  # end位置加入图示中
    # bgraghic.set_title('传感器2_未去基线')

    # save fig
    # 设置新文件的名称
    new_file_name = EQ_Info + '.jpg'
    new_file_name1 = EQ_Info + '.svg'
    # 获取默认路径
    default_path = os.getcwd()
    # 完整的新文件夹路径
    full_folder_path = os.path.join(default_path, 'figureall')
    # 创建新文件夹
    if not os.path.exists(full_folder_path):
        os.makedirs(full_folder_path)
    # 完整的新文件路径
    full_file_path = os.path.join(full_folder_path, new_file_name)
    full_file_path1 = os.path.join(full_folder_path, new_file_name1)
    # 保存图片
    plt.savefig(full_file_path)
    plt.savefig(full_file_path1)
